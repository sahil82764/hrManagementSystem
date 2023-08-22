from tkinter import *
from tkinter import ttk
# from PIL import ImageTk, Image
from tkinter import messagebox
from tkinter import filedialog
from util import util
from openpyxl import load_workbook
import billView
import pandas as pd
import dashboard
import generateBill


class WageView:
    def __init__(self, window, billPath, current_month_claimed_mandays, last_month_claimed_mandays, current_month_active_mandays, lastMonth, billMonth, lastYear, billYear):
        self.window = window
        window.geometry("1366x768")
        window.resizable(0, 0)
        self.window.state('zoomed')
        window.title("WAGE DETAILS")
        self.txt = "WAGE RATE"
        self.color = ["#4f4e4d", "#f29844", "red2"]
        self.heading = Label(self.window, text=self.txt, font=('yu gothic ui', 30, "bold"), bg="white",
                             fg='black',
                             bd=5,
                             relief=FLAT)
        self.heading.grid(row=0, column=0, columnspan=3)

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview.Heading", font=('yu gothic ui', 10, "bold"), foreground="black",
                        background="#108cff")
        
        # ============== INPUT VARIABLES =============================
        self.billPath = billPath
        self. current_month_claimed_mandays = current_month_claimed_mandays
        self.last_month_claimed_mandays = last_month_claimed_mandays
        self.current_month_active_mandays = current_month_active_mandays
        self.lastMonth = lastMonth
        self.billMonth = billMonth
        self.lastYear = lastYear
        self.billYear = billYear

        self.filePath = None

        self.wageRateFile = StringVar()

        # ==========================================================================

        self.dsm_pm = StringVar()
        self.tech_pm = StringVar()
        self.mgr_pm = StringVar()
        self.dsm_fh_pm = StringVar()
        self.tech_fh_pm = StringVar()
        self.mgr_fh_pm = StringVar()
        self.dsm_nh_pm = StringVar()
        self.tech_nh_pm = StringVar()
        self.mgr_nh_pm = StringVar()
        self.dsm_cl_pm = StringVar()
        self.tech_cl_pm = StringVar()
        self.mgr_cl_pm = StringVar()
        self.dsm_fs_pm = StringVar()
        self.tech_fs_pm = StringVar()
        self.mgr_fs_pm = StringVar()

        self.dsm_allw_pm = StringVar()
        self.tech_allw_pm = StringVar()
        self.mgr_allw_pm = StringVar()
        self.dsm_fh_allw_pm = StringVar()
        self.tech_fh_allw_pm = StringVar()
        self.mgr_fh_allw_pm = StringVar()
        self.dsm_nh_allw_pm = StringVar()
        self.tech_nh_allw_pm = StringVar()
        self.mgr_nh_allw_pm = StringVar()
        self.dsm_cl_allw_pm = StringVar()
        self.tech_cl_allw_pm = StringVar()
        self.mgr_cl_allw_pm = StringVar()
        self.dsm_fs_allw_pm = StringVar()
        self.tech_fs_allw_pm = StringVar()
        self.mgr_fs_allw_pm = StringVar()

        self.dsm_cm = StringVar()
        self.tech_cm = StringVar()
        self.mgr_cm = StringVar()
        self.dsm_fh_cm = StringVar()
        self.tech_fh_cm = StringVar()
        self.mgr_fh_cm = StringVar()
        self.dsm_nh_cm = StringVar()
        self.tech_nh_cm = StringVar()
        self.mgr_nh_cm = StringVar()
        self.dsm_cl_cm = StringVar()
        self.tech_cl_cm = StringVar()
        self.mgr_cl_cm = StringVar()
        self.dsm_fs_cm = StringVar()
        self.tech_fs_cm = StringVar()
        self.mgr_fs_cm = StringVar()

        self.dsm_allw_cm = StringVar()
        self.tech_allw_cm = StringVar()
        self.mgr_allw_cm = StringVar()
        self.dsm_fh_allw_cm = StringVar()
        self.tech_fh_allw_cm = StringVar()
        self.mgr_fh_allw_cm = StringVar()
        self.dsm_nh_allw_cm = StringVar()
        self.tech_nh_allw_cm = StringVar()
        self.mgr_nh_allw_cm = StringVar()
        self.dsm_cl_allw_cm = StringVar()
        self.tech_cl_allw_cm = StringVar()
        self.mgr_cl_allw_cm = StringVar()
        self.dsm_fs_allw_cm = StringVar()
        self.tech_fs_allw_cm = StringVar()
        self.mgr_fs_allw_cm = StringVar()

        # ==========================================================================

        self.get_wage_entries()

        # =============== HEADER LABELS =================================

        self.particular_label = Label(self.window, text="PARTICULAR")
        self.particular_label.grid(row=1, column=0, padx=(50, 10), pady=10)

        self.previousMonthWage_label = Label(self.window, text="PREVIOUS MONTH WAGE")
        self.previousMonthWage_label.grid(row=1, column=1, padx=(50, 10), pady=10)

        self.currentMonthWage_label = Label(self.window, text="CURRENT MONTH WAGE")
        self.currentMonthWage_label.grid(row=1, column=2, padx=(50, 10), pady=10)

        # =============== PARTICULAR LABELS =================================

        self.dsm_label = Label(self.window, text="DSM")
        self.dsm_label.grid(row=2, column=0, padx=(50, 10))
        self.dsm_allw_label = Label(self.window, text="DSM Addl.Allowance")
        self.dsm_allw_label.grid(row=3, column=0, padx=(50, 10))

        self.tech_label = Label(self.window, text="TECH")
        self.tech_label.grid(row=4, column=0, padx=(50, 10))
        self.tech_allw_label = Label(self.window, text="TECH Addl.Allowance")
        self.tech_allw_label.grid(row=5, column=0, padx=(50, 10))

        self.mgr_label = Label(self.window, text="MANAGER")
        self.mgr_label.grid(row=6, column=0, padx=(50, 10))
        self.mgr_allw_label = Label(self.window, text="MANAGER Addl.Allowance")
        self.mgr_allw_label.grid(row=7, column=0, padx=(50, 10))

        self.dsm_fh_label = Label(self.window, text="DSM FH")
        self.dsm_fh_label.grid(row=8, column=0, padx=(50, 10))
        self.dsm_fh_allw_label = Label(self.window, text="DSM FH Addl.Allowance")
        self.dsm_fh_allw_label.grid(row=9, column=0, padx=(50, 10))

        self.tech_fh_label = Label(self.window, text="TECH FH")
        self.tech_fh_label.grid(row=10, column=0, padx=(50, 10))
        self.tech_fh_allw_label = Label(self.window, text="TECH FH Addl.Allowance")
        self.tech_fh_allw_label.grid(row=11, column=0, padx=(50, 10))

        self.mgr_fh_label = Label(self.window, text="MANAGER FH")
        self.mgr_fh_label.grid(row=12, column=0, padx=(50, 10))
        self.mgr_fh_allw_label = Label(self.window, text="MANAGER FH Addl.Allowance")
        self.mgr_fh_allw_label.grid(row=13, column=0, padx=(50, 10))

        self.dsm_nh_label = Label(self.window, text="DSM NH")
        self.dsm_nh_label.grid(row=14, column=0, padx=(50, 10))
        self.dsm_nh_allw_label = Label(self.window, text="DSM NH Addl.Allowance")
        self.dsm_nh_allw_label.grid(row=15, column=0, padx=(50, 10))

        self.tech_nh_label = Label(self.window, text="TECH NH")
        self.tech_nh_label.grid(row=16, column=0, padx=(50, 10))
        self.tech_nh_allw_label = Label(self.window, text="TECH NH Addl.Allowance")
        self.tech_nh_allw_label.grid(row=17, column=0, padx=(50, 10))

        self.mgr_nh_label = Label(self.window, text="MANAGER NH")
        self.mgr_nh_label.grid(row=18, column=0, padx=(50, 10))
        self.mgr_nh_allw_label = Label(self.window, text="MANAGER NH Addl.Allowance")
        self.mgr_nh_allw_label.grid(row=19, column=0, padx=(50, 10))

        self.dsm_cl_label = Label(self.window, text="DSM CL")
        self.dsm_cl_label.grid(row=20, column=0, padx=(50, 10))
        self.dsm_cl_allw_label = Label(self.window, text="DSM CL Addl.Allowance")
        self.dsm_cl_allw_label.grid(row=21, column=0, padx=(50, 10))

        self.tech_cl_label = Label(self.window, text="TECH CL")
        self.tech_cl_label.grid(row=22, column=0, padx=(50, 10))
        self.tech_cl_allw_label = Label(self.window, text="TECH CL Addl.Allowance")
        self.tech_cl_allw_label.grid(row=23, column=0, padx=(50, 10))

        self.mgr_cl_label = Label(self.window, text="MANAGER CL")
        self.mgr_cl_label.grid(row=24, column=0, padx=(50, 10))
        self.mgr_cl_allw_label = Label(self.window, text="MANAGER CL Addl.Allowance")
        self.mgr_cl_allw_label.grid(row=25, column=0, padx=(50, 10))

        self.dsm_ft_label = Label(self.window, text="DSM FT")
        self.dsm_ft_label.grid(row=26, column=0, padx=(50, 10))
        self.dsm_ft_allw_label = Label(self.window, text="DSM FT Addl.Allowance")
        self.dsm_ft_allw_label.grid(row=27, column=0, padx=(50, 10))

        self.tech_ft_label = Label(self.window, text="TECH FT")
        self.tech_ft_label.grid(row=28, column=0, padx=(50, 10))
        self.tech_ft_allw_label = Label(self.window, text="TECH FT Addl.Allowance")
        self.tech_ft_allw_label.grid(row=29, column=0, padx=(50, 10))

        self.mgr_ft_label = Label(self.window, text="MANAGER FT")
        self.mgr_ft_label.grid(row=30, column=0, padx=(50, 10))
        self.mgr_ft_allw_label = Label(self.window, text="MANAGER FT Addl.Allowance")
        self.mgr_ft_allw_label.grid(row=31, column=0, padx=(50, 10))


        # =============== PREVIOUS MONTH ENTRIES =================================

        self.dsm_pm_entry = Entry(self.window, width=20, textvariable=self.dsm_pm, state='readonly')
        self.dsm_pm_entry.grid(row=2, column=1, padx=(50, 10))
        self.dsm_allw_pm_entry = Entry(self.window, width=20, textvariable=self.dsm_allw_pm, state='readonly')
        self.dsm_allw_pm_entry.grid(row=3, column=1, padx=(50, 10))

        self.tech_pm_entry = Entry(self.window, width=20, textvariable=self.tech_pm, state='readonly')
        self.tech_pm_entry.grid(row=4, column=1, padx=(50, 10))
        self.tech_allw_pm_entry = Entry(self.window, width=20, textvariable=self.tech_allw_pm, state='readonly')
        self.tech_allw_pm_entry.grid(row=5, column=1, padx=(50, 10))

        self.mgr_pm_entry = Entry(self.window, width=20, textvariable=self.mgr_pm, state='readonly')
        self.mgr_pm_entry.grid(row=6, column=1, padx=(50, 10))
        self.mgr_allw_pm_entry = Entry(self.window, width=20, textvariable=self.mgr_allw_pm, state='readonly')
        self.mgr_allw_pm_entry.grid(row=7, column=1, padx=(50, 10))

        self.dsm_fh_pm_entry = Entry(self.window, width=20, textvariable=self.dsm_fh_pm, state='readonly')
        self.dsm_fh_pm_entry.grid(row=8, column=1, padx=(50, 10))
        self.dsm_fh_allw_pm_entry = Entry(self.window, width=20, textvariable=self.dsm_fh_allw_pm, state='readonly')
        self.dsm_fh_allw_pm_entry.grid(row=9, column=1, padx=(50, 10))

        self.tech_fh_pm_entry = Entry(self.window, width=20, textvariable=self.tech_fh_pm, state='readonly')
        self.tech_fh_pm_entry.grid(row=10, column=1, padx=(50, 10))
        self.tech_fh_allw_pm_entry = Entry(self.window, width=20, textvariable=self.tech_fh_allw_pm, state='readonly')
        self.tech_fh_allw_pm_entry.grid(row=11, column=1, padx=(50, 10))

        self.mgr_fh_pm_entry = Entry(self.window, width=20, textvariable=self.mgr_fh_pm, state='readonly')
        self.mgr_fh_pm_entry.grid(row=12, column=1, padx=(50, 10))
        self.mgr_fh_allw_pm_entry = Entry(self.window, width=20, textvariable=self.mgr_fh_allw_pm, state='readonly')
        self.mgr_fh_allw_pm_entry.grid(row=13, column=1, padx=(50, 10))

        self.dsm_nh_pm_entry = Entry(self.window, width=20, textvariable=self.dsm_nh_pm, state='readonly')
        self.dsm_nh_pm_entry.grid(row=14, column=1, padx=(50, 10))
        self.dsm_nh_allw_pm_entry = Entry(self.window, width=20, textvariable=self.dsm_nh_allw_pm, state='readonly')
        self.dsm_nh_allw_pm_entry.grid(row=15, column=1, padx=(50, 10))

        self.tech_nh_pm_entry = Entry(self.window, width=20, textvariable=self.tech_nh_pm, state='readonly')
        self.tech_nh_pm_entry.grid(row=16, column=1, padx=(50, 10))
        self.tech_nh_allw_pm_entry = Entry(self.window, width=20, textvariable=self.tech_nh_allw_pm, state='readonly')
        self.tech_nh_allw_pm_entry.grid(row=17, column=1, padx=(50, 10))

        self.mgr_nh_pm_entry = Entry(self.window, width=20, textvariable=self.mgr_nh_pm, state='readonly')
        self.mgr_nh_pm_entry.grid(row=18, column=1, padx=(50, 10))
        self.mgr_nh_allw_pm_entry = Entry(self.window, width=20, textvariable=self.mgr_nh_allw_pm, state='readonly')
        self.mgr_nh_allw_pm_entry.grid(row=19, column=1, padx=(50, 10))

        self.dsm_cl_pm_entry = Entry(self.window, width=20, textvariable=self.dsm_cl_pm, state='readonly')
        self.dsm_cl_pm_entry.grid(row=20, column=1, padx=(50, 10))
        self.dsm_cl_allw_pm_entry = Entry(self.window, width=20, textvariable=self.dsm_cl_allw_pm, state='readonly')
        self.dsm_cl_allw_pm_entry.grid(row=21, column=1, padx=(50, 10))

        self.tech_cl_pm_entry = Entry(self.window, width=20, textvariable=self.tech_cl_pm, state='readonly')
        self.tech_cl_pm_entry.grid(row=22, column=1, padx=(50, 10))
        self.tech_cl_allw_pm_entry = Entry(self.window, width=20, textvariable=self.tech_cl_allw_pm, state='readonly')
        self.tech_cl_allw_pm_entry.grid(row=23, column=1, padx=(50, 10))

        self.mgr_cl_pm_entry = Entry(self.window, width=20, textvariable=self.mgr_cl_pm, state='readonly')
        self.mgr_cl_pm_entry.grid(row=24, column=1, padx=(50, 10))
        self.mgr_cl_allw_pm_entry = Entry(self.window, width=20, textvariable=self.mgr_cl_allw_pm, state='readonly')
        self.mgr_cl_allw_pm_entry.grid(row=25, column=1, padx=(50, 10))

        self.dsm_ft_pm_entry = Entry(self.window, width=20, textvariable=self.dsm_fs_pm, state='readonly')
        self.dsm_ft_pm_entry.grid(row=26, column=1, padx=(50, 10))
        self.dsm_ft_allw_pm_entry = Entry(self.window, width=20, textvariable=self.dsm_fs_allw_pm, state='readonly')
        self.dsm_ft_allw_pm_entry.grid(row=27, column=1, padx=(50, 10))

        self.tech_ft_pm_entry = Entry(self.window, width=20, textvariable=self.tech_fs_pm, state='readonly')
        self.tech_ft_pm_entry.grid(row=28, column=1, padx=(50, 10))
        self.tech_ft_allw_pm_entry = Entry(self.window, width=20, textvariable=self.tech_fs_allw_pm, state='readonly')
        self.tech_ft_allw_pm_entry.grid(row=29, column=1, padx=(50, 10))

        self.mgr_ft_pm_entry = Entry(self.window, width=20, textvariable=self.mgr_fs_pm, state='readonly')
        self.mgr_ft_pm_entry.grid(row=30, column=1, padx=(50, 10))
        self.mgr_ft_allw_pm_entry = Entry(self.window, width=20, textvariable=self.mgr_fs_allw_pm, state='readonly')
        self.mgr_ft_allw_pm_entry.grid(row=31, column=1, padx=(50, 10))

        # =============== CURRENT MONTH ENTRIES =================================

        self.dsm_cm_entry = Entry(self.window, width=20, textvariable=self.dsm_cm, state='readonly')
        self.dsm_cm_entry.grid(row=2, column=2, padx=(50, 10))
        self.dsm_allw_cm_entry = Entry(self.window, width=20, textvariable=self.dsm_allw_cm, state='readonly')
        self.dsm_allw_cm_entry.grid(row=3, column=2, padx=(50, 10))

        self.tech_cm_entry = Entry(self.window, width=20, textvariable=self.tech_cm, state='readonly')
        self.tech_cm_entry.grid(row=4, column=2, padx=(50, 10))
        self.tech_allw_cm_entry = Entry(self.window, width=20, textvariable=self.tech_allw_cm, state='readonly')
        self.tech_allw_cm_entry.grid(row=5, column=2, padx=(50, 10))

        self.mgr_cm_entry = Entry(self.window, width=20, textvariable=self.mgr_cm, state='readonly')
        self.mgr_cm_entry.grid(row=6, column=2, padx=(50, 10))
        self.mgr_allw_cm_entry = Entry(self.window, width=20, textvariable=self.mgr_allw_cm, state='readonly')
        self.mgr_allw_cm_entry.grid(row=7, column=2, padx=(50, 10))

        self.dsm_fh_cm_entry = Entry(self.window, width=20, textvariable=self.dsm_fh_cm, state='readonly')
        self.dsm_fh_cm_entry.grid(row=8, column=2, padx=(50, 10))
        self.dsm_fh_allw_cm_entry = Entry(self.window, width=20, textvariable=self.dsm_fh_allw_cm, state='readonly')
        self.dsm_fh_allw_cm_entry.grid(row=9, column=2, padx=(50, 10))

        self.tech_fh_cm_entry = Entry(self.window, width=20, textvariable=self.tech_fh_cm, state='readonly')
        self.tech_fh_cm_entry.grid(row=10, column=2, padx=(50, 10))
        self.tech_fh_allw_cm_entry = Entry(self.window, width=20, textvariable=self.tech_fh_allw_cm, state='readonly')
        self.tech_fh_allw_cm_entry.grid(row=11, column=2, padx=(50, 10))

        self.mgr_fh_cm_entry = Entry(self.window, width=20, textvariable=self.mgr_fh_cm, state='readonly')
        self.mgr_fh_cm_entry.grid(row=12, column=2, padx=(50, 10))
        self.mgr_fh_allw_cm_entry = Entry(self.window, width=20, textvariable=self.mgr_fh_allw_cm, state='readonly')
        self.mgr_fh_allw_cm_entry.grid(row=13, column=2, padx=(50, 10))

        self.dsm_nh_cm_entry = Entry(self.window, width=20, textvariable=self.dsm_nh_cm, state='readonly')
        self.dsm_nh_cm_entry.grid(row=14, column=2, padx=(50, 10))
        self.dsm_nh_allw_cm_entry = Entry(self.window, width=20, textvariable=self.dsm_nh_allw_cm, state='readonly')
        self.dsm_nh_allw_cm_entry.grid(row=15, column=2, padx=(50, 10))

        self.tech_nh_cm_entry = Entry(self.window, width=20, textvariable=self.tech_nh_cm, state='readonly')
        self.tech_nh_cm_entry.grid(row=16, column=2, padx=(50, 10))
        self.tech_nh_allw_cm_entry = Entry(self.window, width=20, textvariable=self.tech_nh_allw_cm, state='readonly')
        self.tech_nh_allw_cm_entry.grid(row=17, column=2, padx=(50, 10))

        self.mgr_nh_cm_entry = Entry(self.window, width=20, textvariable=self.mgr_nh_cm, state='readonly')
        self.mgr_nh_cm_entry.grid(row=18, column=2, padx=(50, 10))
        self.mgr_nh_allw_cm_entry = Entry(self.window, width=20, textvariable=self.mgr_nh_allw_cm, state='readonly')
        self.mgr_nh_allw_cm_entry.grid(row=19, column=2, padx=(50, 10))

        self.dsm_cl_cm_entry = Entry(self.window, width=20, textvariable=self.dsm_cl_cm, state='readonly')
        self.dsm_cl_cm_entry.grid(row=20, column=2, padx=(50, 10))
        self.dsm_cl_allw_cm_entry = Entry(self.window, width=20, textvariable=self.dsm_cl_allw_cm, state='readonly')
        self.dsm_cl_allw_cm_entry.grid(row=21, column=2, padx=(50, 10))

        self.tech_cl_cm_entry = Entry(self.window, width=20, textvariable=self.tech_cl_cm, state='readonly')
        self.tech_cl_cm_entry.grid(row=22, column=2, padx=(50, 10))
        self.tech_cl_allw_cm_entry = Entry(self.window, width=20, textvariable=self.tech_cl_allw_cm, state='readonly')
        self.tech_cl_allw_cm_entry.grid(row=23, column=2, padx=(50, 10))

        self.mgr_cl_cm_entry = Entry(self.window, width=20, textvariable=self.mgr_cl_cm, state='readonly')
        self.mgr_cl_cm_entry.grid(row=24, column=2, padx=(50, 10))
        self.mgr_cl_allw_cm_entry = Entry(self.window, width=20, textvariable=self.mgr_cl_allw_cm, state='readonly')
        self.mgr_cl_allw_cm_entry.grid(row=25, column=2, padx=(50, 10))

        self.dsm_ft_cm_entry = Entry(self.window, width=20, textvariable=self.dsm_fs_cm, state='readonly')
        self.dsm_ft_cm_entry.grid(row=26, column=2, padx=(50, 10))
        self.dsm_ft_allw_cm_entry = Entry(self.window, width=20, textvariable=self.dsm_fs_allw_cm, state='readonly')
        self.dsm_ft_allw_cm_entry.grid(row=27, column=2, padx=(50, 10))

        self.tech_ft_cm_entry = Entry(self.window, width=20, textvariable=self.tech_fs_cm, state='readonly')
        self.tech_ft_cm_entry.grid(row=28, column=2, padx=(50, 10))
        self.tech_ft_allw_cm_entry = Entry(self.window, width=20, textvariable=self.tech_fs_allw_cm, state='readonly')
        self.tech_ft_allw_cm_entry.grid(row=29, column=2, padx=(50, 10))

        self.mgr_ft_cm_entry = Entry(self.window, width=20, textvariable=self.mgr_fs_cm, state='readonly')
        self.mgr_ft_cm_entry.grid(row=30, column=2, padx=(50, 10))
        self.mgr_ft_allw_cm_entry = Entry(self.window, width=20, textvariable=self.mgr_fs_allw_cm, state='readonly')
        self.mgr_ft_allw_cm_entry.grid(row=31, column=2, padx=(50, 10))

        
        # Create the "Change Entries" button
        self.change_btn = Button(self.window, text="SAVE ENTRIES", command=lambda: self.toggleEntry())
        self.change_btn.grid(row=32, column=0, columnspan=2, padx=(50, 10), pady=4, sticky="ew")

        # # Create the "Save Entries" button
        self.save_btn = Button(self.window, text="SAVE ENTRIES", command=lambda: self.saveEntry(), state='disabled')
        self.save_btn.grid(row=32, column=2, columnspan=3, padx=(50, 10), pady=4, sticky="ew")

        # # Create the "Generate Bill" button
        self.bill_btn = Button(self.window, text="GENERATE BILL", command=lambda: self.generate_bill())
        self.bill_btn.grid(row=33, column=1, columnspan=3, padx=(50, 10), pady=4, sticky="ew")

        # # Create the "Back" button
        self.back_btn = Button(self.window, text="BACK", command=lambda: self.back_operation())
        self.back_btn.grid(row=33, column=4, columnspan=2, padx=(50, 10), pady=4, sticky="ew")

        for i in range(2, 32):
            self.label = Label(self.window, text=" || ")
            self.label.grid(row=i, column=4, padx=(50, 10))

        self.wageRate_entry = Entry(self.window, textvariable=self.wageRateFile, width=60)
        self.wageRate_entry.grid(row=14, column=5, columnspan=5, padx=(50, 10), sticky="ew")

        self.upload_btn = Button(self.window, text="UPLOAD WAGE RATE", command=lambda: self.upload_wage_rate(), width= 50)
        self.upload_btn.grid(row=16, column=5, padx=(50, 10))


    def get_wage_entries(self):

        self.wageWorkbook = load_workbook(util.get_wage_template())
        self.wage_sheet = self.wageWorkbook.active

        # ============== PREVIOUS MONTH WAGE VARIABLES =============================        
        self.dsm_pm.set(self.wage_sheet['B2'].value)        
        self.tech_pm.set(self.wage_sheet['B4'].value)        
        self.mgr_pm.set(self.wage_sheet['B6'].value)        
        
        self.dsm_fh_pm.set(self.wage_sheet['B8'].value)        
        self.tech_fh_pm.set(self.wage_sheet['B10'].value)        
        self.mgr_fh_pm.set(self.wage_sheet['B12'].value)
        
        self.dsm_nh_pm.set(self.wage_sheet['B14'].value)        
        self.tech_nh_pm.set(self.wage_sheet['B16'].value)        
        self.mgr_nh_pm.set(self.wage_sheet['B18'].value)        
        
        self.dsm_cl_pm.set(self.wage_sheet['B20'].value)        
        self.tech_cl_pm.set(self.wage_sheet['B22'].value)        
        self.mgr_cl_pm.set(self.wage_sheet['B24'].value)
        
        self.dsm_fs_pm.set(self.wage_sheet['B26'].value)        
        self.tech_fs_pm.set(self.wage_sheet['B28'].value)        
        self.mgr_fs_pm.set(self.wage_sheet['B30'].value)

        self.dsm_allw_pm.set(self.wage_sheet['B3'].value)        
        self.tech_allw_pm.set(self.wage_sheet['B5'].value)        
        self.mgr_allw_pm.set(self.wage_sheet['B7'].value)        
        
        self.dsm_fh_allw_pm.set(self.wage_sheet['B9'].value)        
        self.tech_fh_allw_pm.set(self.wage_sheet['B11'].value)        
        self.mgr_fh_allw_pm.set(self.wage_sheet['B13'].value)
        
        self.dsm_nh_allw_pm.set(self.wage_sheet['B15'].value)        
        self.tech_nh_allw_pm.set(self.wage_sheet['B17'].value)        
        self.mgr_nh_allw_pm.set(self.wage_sheet['B19'].value)        
        
        self.dsm_cl_allw_pm.set(self.wage_sheet['B21'].value)        
        self.tech_cl_allw_pm.set(self.wage_sheet['B23'].value)        
        self.mgr_cl_allw_pm.set(self.wage_sheet['B25'].value)
        
        self.dsm_fs_allw_pm.set(self.wage_sheet['B27'].value)        
        self.tech_fs_allw_pm.set(self.wage_sheet['B29'].value)        
        self.mgr_fs_allw_pm.set(self.wage_sheet['B31'].value)

        # ============== CURRENT MONTH WAGE VARIABLES =============================        
        self.dsm_cm.set(self.wage_sheet['C2'].value)        
        self.tech_cm.set(self.wage_sheet['C4'].value)        
        self.mgr_cm.set(self.wage_sheet['C6'].value)        
        
        self.dsm_fh_cm.set(self.wage_sheet['C8'].value)        
        self.tech_fh_cm.set(self.wage_sheet['C10'].value)        
        self.mgr_fh_cm.set(self.wage_sheet['C12'].value)
        
        self.dsm_nh_cm.set(self.wage_sheet['C14'].value)        
        self.tech_nh_cm.set(self.wage_sheet['C16'].value)        
        self.mgr_nh_cm.set(self.wage_sheet['C18'].value)        
        
        self.dsm_cl_cm.set(self.wage_sheet['C20'].value)        
        self.tech_cl_cm.set(self.wage_sheet['C22'].value)        
        self.mgr_cl_cm.set(self.wage_sheet['C24'].value)
        
        self.dsm_fs_cm.set(self.wage_sheet['C26'].value)        
        self.tech_fs_cm.set(self.wage_sheet['C28'].value)        
        self.mgr_fs_cm.set(self.wage_sheet['C30'].value)

        self.dsm_allw_cm.set(self.wage_sheet['C3'].value)        
        self.tech_allw_cm.set(self.wage_sheet['C5'].value)        
        self.mgr_allw_cm.set(self.wage_sheet['C7'].value)        
        
        self.dsm_fh_allw_cm.set(self.wage_sheet['C9'].value)        
        self.tech_fh_allw_cm.set(self.wage_sheet['C11'].value)        
        self.mgr_fh_allw_cm.set(self.wage_sheet['C13'].value)
        
        self.dsm_nh_allw_cm.set(self.wage_sheet['C15'].value)        
        self.tech_nh_allw_cm.set(self.wage_sheet['C17'].value)        
        self.mgr_nh_allw_cm.set(self.wage_sheet['C19'].value)        
        
        self.dsm_cl_allw_cm.set(self.wage_sheet['C21'].value)        
        self.tech_cl_allw_cm.set(self.wage_sheet['C23'].value)        
        self.mgr_cl_allw_cm.set(self.wage_sheet['C25'].value)
        
        self.dsm_fs_allw_cm.set(self.wage_sheet['C27'].value)        
        self.tech_fs_allw_cm.set(self.wage_sheet['C29'].value)        
        self.mgr_fs_allw_cm.set(self.wage_sheet['C31'].value)


        # ==================== SAVING WORKBOOK =================
        self.wageWorkbook.save(util.get_wage_template())


    def toggleEntry(self):

        self.dsm_pm_entry.config(state='normal')
        self.tech_pm_entry.config(state='normal')
        self.mgr_pm_entry.config(state='normal')
        self.dsm_fh_pm_entry.config(state='normal')
        self.tech_fh_pm_entry.config(state='normal')
        self.mgr_fh_pm_entry.config(state='normal')
        self.dsm_nh_pm_entry.config(state='normal')
        self.tech_nh_pm_entry.config(state='normal')
        self.mgr_nh_pm_entry.config(state='normal')
        self.dsm_cl_pm_entry.config(state='normal')
        self.tech_cl_pm_entry.config(state='normal')
        self.mgr_cl_pm_entry.config(state='normal')
        self.dsm_ft_pm_entry.config(state='normal')
        self.tech_ft_pm_entry.config(state='normal')
        self.mgr_ft_pm_entry.config(state='normal')

        self.dsm_allw_pm_entry.config(state='normal')
        self.tech_allw_pm_entry.config(state='normal')
        self.mgr_allw_pm_entry.config(state='normal')
        self.dsm_fh_allw_pm_entry.config(state='normal')
        self.tech_fh_allw_pm_entry.config(state='normal')
        self.mgr_fh_allw_pm_entry.config(state='normal')
        self.dsm_nh_allw_pm_entry.config(state='normal')
        self.tech_nh_allw_pm_entry.config(state='normal')
        self.mgr_nh_allw_pm_entry.config(state='normal')
        self.dsm_cl_allw_pm_entry.config(state='normal')
        self.tech_cl_allw_pm_entry.config(state='normal')
        self.mgr_cl_allw_pm_entry.config(state='normal')
        self.dsm_ft_allw_pm_entry.config(state='normal')
        self.tech_ft_allw_pm_entry.config(state='normal')
        self.mgr_ft_allw_pm_entry.config(state='normal')

        self.dsm_cm_entry.config(state='normal')
        self.tech_cm_entry.config(state='normal')
        self.mgr_cm_entry.config(state='normal')
        self.dsm_fh_cm_entry.config(state='normal')
        self.tech_fh_cm_entry.config(state='normal')
        self.mgr_fh_cm_entry.config(state='normal')
        self.dsm_nh_cm_entry.config(state='normal')
        self.tech_nh_cm_entry.config(state='normal')
        self.mgr_nh_cm_entry.config(state='normal')
        self.dsm_cl_cm_entry.config(state='normal')
        self.tech_cl_cm_entry.config(state='normal')
        self.mgr_cl_cm_entry.config(state='normal')
        self.dsm_ft_cm_entry.config(state='normal')
        self.tech_ft_cm_entry.config(state='normal')
        self.mgr_ft_cm_entry.config(state='normal')

        self.dsm_allw_cm_entry.config(state='normal')
        self.tech_allw_cm_entry.config(state='normal')
        self.mgr_allw_cm_entry.config(state='normal')
        self.dsm_fh_allw_cm_entry.config(state='normal')
        self.tech_fh_allw_cm_entry.config(state='normal')
        self.mgr_fh_allw_cm_entry.config(state='normal')
        self.dsm_nh_allw_cm_entry.config(state='normal')
        self.tech_nh_allw_cm_entry.config(state='normal')
        self.mgr_nh_allw_cm_entry.config(state='normal')
        self.dsm_cl_allw_cm_entry.config(state='normal')
        self.tech_cl_allw_cm_entry.config(state='normal')
        self.mgr_cl_allw_cm_entry.config(state='normal')
        self.dsm_ft_allw_cm_entry.config(state='normal')
        self.tech_ft_allw_cm_entry.config(state='normal')
        self.mgr_ft_allw_cm_entry.config(state='normal')

        # Toggle the button text between "Change" and "Save"
        self.change_btn.config(state='disabled')
        self.save_btn.config(state='normal')


    def saveEntry(self):        
        if (

            self.dsm_pm.get() and
            self.tech_pm.get() and
            self.mgr_pm.get() and
            self.dsm_fh_pm.get() and
            self.tech_fh_pm.get() and
            self.mgr_fh_pm.get() and
            self.dsm_nh_pm.get() and
            self.tech_nh_pm.get() and
            self.mgr_nh_pm.get() and
            self.dsm_cl_pm.get() and
            self.tech_cl_pm.get() and
            self.mgr_cl_pm.get() and
            self.dsm_fs_pm.get() and
            self.tech_fs_pm.get() and
            self.mgr_fs_pm.get() and
            self.dsm_cm.get() and
            self.tech_cm.get() and
            self.mgr_cm.get() and
            self.dsm_fh_cm.get() and
            self.tech_fh_cm.get() and
            self.mgr_fh_cm.get() and
            self.dsm_nh_cm.get() and
            self.tech_nh_cm.get() and
            self.mgr_nh_cm.get() and
            self.dsm_cl_cm.get() and
            self.tech_cl_cm.get() and
            self.mgr_cl_cm.get() and
            self.dsm_fs_cm.get() and
            self.tech_fs_cm.get() and
            self.mgr_fs_cm.get()
            
        ):
            
            self.wageWorkbook = load_workbook(util.get_wage_template())
            self.wage_sheet = self.wageWorkbook.active
            
            self.wage_sheet['B2'] = int(self.dsm_pm.get())
            self.wage_sheet['B4'] = int(self.tech_pm.get())
            self.wage_sheet['B6'] = int(self.mgr_pm.get())
            self.wage_sheet['B8'] = int(self.dsm_fh_pm.get())
            self.wage_sheet['B10'] = int(self.tech_fh_pm.get())
            self.wage_sheet['B12'] = int(self.mgr_fh_pm.get())
            self.wage_sheet['B14'] = int(self.dsm_nh_pm.get())
            self.wage_sheet['B16'] = int(self.tech_nh_pm.get())
            self.wage_sheet['B18'] = int(self.mgr_nh_pm.get())
            self.wage_sheet['B20'] = int(self.dsm_cl_pm.get())
            self.wage_sheet['B22'] = int(self.tech_cl_pm.get())
            self.wage_sheet['B24'] = int(self.mgr_cl_pm.get())
            self.wage_sheet['B26'] = int(self.dsm_fs_pm.get())
            self.wage_sheet['B28'] = int(self.tech_fs_pm.get())
            self.wage_sheet['B30'] = int(self.mgr_fs_pm.get())

            self.wage_sheet['B3'] = int(self.dsm_allw_pm.get())
            self.wage_sheet['B5'] = int(self.tech_allw_pm.get())
            self.wage_sheet['B7'] = int(self.mgr_allw_pm.get())
            self.wage_sheet['B9'] = int(self.dsm_fh_allw_pm.get())
            self.wage_sheet['B11'] = int(self.tech_fh_allw_pm.get())
            self.wage_sheet['B13'] = int(self.mgr_fh_allw_pm.get())
            self.wage_sheet['B15'] = int(self.dsm_nh_allw_pm.get())
            self.wage_sheet['B17'] = int(self.tech_nh_allw_pm.get())
            self.wage_sheet['B19'] = int(self.mgr_nh_allw_pm.get())
            self.wage_sheet['B21'] = int(self.dsm_cl_allw_pm.get())
            self.wage_sheet['B23'] = int(self.tech_cl_allw_pm.get())
            self.wage_sheet['B25'] = int(self.mgr_cl_allw_pm.get())
            self.wage_sheet['B27'] = int(self.dsm_fs_allw_pm.get())
            self.wage_sheet['B29'] = int(self.tech_fs_allw_pm.get())
            self.wage_sheet['B31'] = int(self.mgr_fs_allw_pm.get())

            self.wage_sheet['C2'] = int(self.dsm_cm.get())
            self.wage_sheet['C4'] = int(self.tech_cm.get())
            self.wage_sheet['C6'] = int(self.mgr_cm.get())
            self.wage_sheet['C8'] = int(self.dsm_fh_cm.get())
            self.wage_sheet['C10'] = int(self.tech_fh_cm.get())
            self.wage_sheet['C12'] = int(self.mgr_fh_cm.get())
            self.wage_sheet['C14'] = int(self.dsm_nh_cm.get())
            self.wage_sheet['C16'] = int(self.tech_nh_cm.get())
            self.wage_sheet['C18'] = int(self.mgr_nh_cm.get())
            self.wage_sheet['C20'] = int(self.dsm_cl_cm.get())
            self.wage_sheet['C22'] = int(self.tech_cl_cm.get())
            self.wage_sheet['C24'] = int(self.mgr_cl_cm.get())
            self.wage_sheet['C26'] = int(self.dsm_fs_cm.get())
            self.wage_sheet['C28'] = int(self.tech_fs_cm.get())
            self.wage_sheet['C30'] = int(self.mgr_fs_cm.get())

            self.wage_sheet['C3'] = int(self.dsm_allw_cm.get())
            self.wage_sheet['C5'] = int(self.tech_allw_cm.get())
            self.wage_sheet['C7'] = int(self.mgr_allw_cm.get())
            self.wage_sheet['C9'] = int(self.dsm_fh_allw_cm.get())
            self.wage_sheet['C11'] = int(self.tech_fh_allw_cm.get())
            self.wage_sheet['C13'] = int(self.mgr_fh_allw_cm.get())
            self.wage_sheet['C15'] = int(self.dsm_nh_allw_cm.get())
            self.wage_sheet['C17'] = int(self.tech_nh_allw_cm.get())
            self.wage_sheet['C19'] = int(self.mgr_nh_allw_cm.get())
            self.wage_sheet['C21'] = int(self.dsm_cl_allw_cm.get())
            self.wage_sheet['C23'] = int(self.tech_cl_allw_cm.get())
            self.wage_sheet['C25'] = int(self.mgr_cl_allw_cm.get())
            self.wage_sheet['C27'] = int(self.dsm_fs_allw_cm.get())
            self.wage_sheet['C29'] = int(self.tech_fs_allw_cm.get())
            self.wage_sheet['C31'] = int(self.mgr_fs_allw_cm.get())


            # ============== SAVE WORKBOOK =================
            self.wageWorkbook.save(util.get_wage_template())


            # ========= Change Config to READONLY ==========

            self.dsm_pm_entry.config(state='readonly')
            self.tech_pm_entry.config(state='readonly')
            self.mgr_pm_entry.config(state='readonly')
            self.dsm_fh_pm_entry.config(state='readonly')
            self.tech_fh_pm_entry.config(state='readonly')
            self.mgr_fh_pm_entry.config(state='readonly')
            self.dsm_nh_pm_entry.config(state='readonly')
            self.tech_nh_pm_entry.config(state='readonly')
            self.mgr_nh_pm_entry.config(state='readonly')
            self.dsm_cl_pm_entry.config(state='readonly')
            self.tech_cl_pm_entry.config(state='readonly')
            self.mgr_cl_pm_entry.config(state='readonly')
            self.dsm_ft_pm_entry.config(state='readonly')
            self.tech_ft_pm_entry.config(state='readonly')
            self.mgr_ft_pm_entry.config(state='readonly')

            self.dsm_allw_pm_entry.config(state='readonly')
            self.tech_allw_pm_entry.config(state='readonly')
            self.mgr_allw_pm_entry.config(state='readonly')
            self.dsm_fh_allw_pm_entry.config(state='readonly')
            self.tech_fh_allw_pm_entry.config(state='readonly')
            self.mgr_fh_allw_pm_entry.config(state='readonly')
            self.dsm_nh_allw_pm_entry.config(state='readonly')
            self.tech_nh_allw_pm_entry.config(state='readonly')
            self.mgr_nh_allw_pm_entry.config(state='readonly')
            self.dsm_cl_allw_pm_entry.config(state='readonly')
            self.tech_cl_allw_pm_entry.config(state='readonly')
            self.mgr_cl_allw_pm_entry.config(state='readonly')
            self.dsm_ft_allw_pm_entry.config(state='readonly')
            self.tech_ft_allw_pm_entry.config(state='readonly')
            self.mgr_ft_allw_pm_entry.config(state='readonly')

            self.dsm_cm_entry.config(state='readonly')
            self.tech_cm_entry.config(state='readonly')
            self.mgr_cm_entry.config(state='readonly')
            self.dsm_fh_cm_entry.config(state='readonly')
            self.tech_fh_cm_entry.config(state='readonly')
            self.mgr_fh_cm_entry.config(state='readonly')
            self.dsm_nh_cm_entry.config(state='readonly')
            self.tech_nh_cm_entry.config(state='readonly')
            self.mgr_nh_cm_entry.config(state='readonly')
            self.dsm_cl_cm_entry.config(state='readonly')
            self.tech_cl_cm_entry.config(state='readonly')
            self.mgr_cl_cm_entry.config(state='readonly')
            self.dsm_ft_cm_entry.config(state='readonly')
            self.tech_ft_cm_entry.config(state='readonly')
            self.mgr_ft_cm_entry.config(state='readonly')

            self.dsm_allw_cm_entry.config(state='readonly')
            self.tech_allw_cm_entry.config(state='readonly')
            self.mgr_allw_cm_entry.config(state='readonly')
            self.dsm_fh_allw_cm_entry.config(state='readonly')
            self.tech_fh_allw_cm_entry.config(state='readonly')
            self.mgr_fh_allw_cm_entry.config(state='readonly')
            self.dsm_nh_allw_cm_entry.config(state='readonly')
            self.tech_nh_allw_cm_entry.config(state='readonly')
            self.mgr_nh_allw_cm_entry.config(state='readonly')
            self.dsm_cl_allw_cm_entry.config(state='readonly')
            self.tech_cl_allw_cm_entry.config(state='readonly')
            self.mgr_cl_allw_cm_entry.config(state='readonly')
            self.dsm_ft_allw_cm_entry.config(state='readonly')
            self.tech_ft_allw_cm_entry.config(state='readonly')
            self.mgr_ft_allw_cm_entry.config(state='readonly')

            self.change_btn.config(state='normal')
            self.save_btn.config(state='disabled')

        else:
            # Display an error message if any field is empty
            messagebox.showerror("Error", "Please fill in all the fields.")


    def upload_wage_rate(self):
        try:
            self.filePath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
            self.wageRate_entry.config(state='normal')
            self.wageRate_entry.delete(0, END)  # Clear previous path, if any
            self.wageRate_entry.insert(END, self.filePath)  # Display the selected path
            self.wageRate_entry.config(state='readonly')
        
        except Exception as e:
            print(e)

    def generate_bill(self):
        try:
            if (

                self.dsm_pm.get() and
                self.tech_pm.get() and
                self.mgr_pm.get() and
                self.dsm_fh_pm.get() and
                self.tech_fh_pm.get() and
                self.mgr_fh_pm.get() and
                self.dsm_nh_pm.get() and
                self.tech_nh_pm.get() and
                self.mgr_nh_pm.get() and
                self.dsm_cl_pm.get() and
                self.tech_cl_pm.get() and
                self.mgr_cl_pm.get() and
                self.dsm_fs_pm.get() and
                self.tech_fs_pm.get() and
                self.mgr_fs_pm.get() and
                self.dsm_cm.get() and
                self.tech_cm.get() and
                self.mgr_cm.get() and
                self.dsm_fh_cm.get() and
                self.tech_fh_cm.get() and
                self.mgr_fh_cm.get() and
                self.dsm_nh_cm.get() and
                self.tech_nh_cm.get() and
                self.mgr_nh_cm.get() and
                self.dsm_cl_cm.get() and
                self.tech_cl_cm.get() and
                self.mgr_cl_cm.get() and
                self.dsm_fs_cm.get() and
                self.tech_fs_cm.get() and
                self.mgr_fs_cm.get()
                
            ):

                self.billWorkbook = load_workbook(self.billPath)
                self.bill_sheet = self.billWorkbook.active
                
                self.bill_sheet['J12'] = int(self.dsm_pm.get())
                self.bill_sheet['J14'] = int(self.tech_pm.get())
                self.bill_sheet['J16'] = int(self.mgr_pm.get())
                self.bill_sheet['J18'] = int(self.dsm_fh_pm.get())
                self.bill_sheet['J20'] = int(self.tech_fh_pm.get())
                self.bill_sheet['J22'] = int(self.mgr_fh_pm.get())
                self.bill_sheet['J24'] = int(self.dsm_nh_pm.get())
                self.bill_sheet['J26'] = int(self.tech_nh_pm.get())
                self.bill_sheet['J28'] = int(self.mgr_nh_pm.get())
                self.bill_sheet['J30'] = int(self.dsm_cl_pm.get())
                self.bill_sheet['J32'] = int(self.tech_cl_pm.get())
                self.bill_sheet['J34'] = int(self.mgr_cl_pm.get())
                self.bill_sheet['J36'] = int(self.dsm_fs_pm.get())
                self.bill_sheet['J38'] = int(self.tech_fs_pm.get())
                self.bill_sheet['J40'] = int(self.mgr_fs_pm.get())

                self.bill_sheet['J13'] = int(self.dsm_allw_pm.get())
                self.bill_sheet['J15'] = int(self.tech_allw_pm.get())
                self.bill_sheet['J17'] = int(self.mgr_allw_pm.get())
                self.bill_sheet['J19'] = int(self.dsm_fh_allw_pm.get())
                self.bill_sheet['J21'] = int(self.tech_fh_allw_pm.get())
                self.bill_sheet['J23'] = int(self.mgr_fh_allw_pm.get())
                self.bill_sheet['J25'] = int(self.dsm_nh_allw_pm.get())
                self.bill_sheet['J27'] = int(self.tech_nh_allw_pm.get())
                self.bill_sheet['J29'] = int(self.mgr_nh_allw_pm.get())
                self.bill_sheet['J31'] = int(self.dsm_cl_allw_pm.get())
                self.bill_sheet['J33'] = int(self.tech_cl_allw_pm.get())
                self.bill_sheet['J35'] = int(self.mgr_cl_allw_pm.get())
                self.bill_sheet['J37'] = int(self.dsm_fs_allw_pm.get())
                self.bill_sheet['J39'] = int(self.tech_fs_allw_pm.get())
                self.bill_sheet['J41'] = int(self.mgr_fs_allw_pm.get())

                self.bill_sheet['K12'] = int(self.dsm_cm.get())
                self.bill_sheet['K14'] = int(self.tech_cm.get())
                self.bill_sheet['K16'] = int(self.mgr_cm.get())
                self.bill_sheet['K18'] = int(self.dsm_fh_cm.get())
                self.bill_sheet['K20'] = int(self.tech_fh_cm.get())
                self.bill_sheet['K22'] = int(self.mgr_fh_cm.get())
                self.bill_sheet['K24'] = int(self.dsm_nh_cm.get())
                self.bill_sheet['K26'] = int(self.tech_nh_cm.get())
                self.bill_sheet['K28'] = int(self.mgr_nh_cm.get())
                self.bill_sheet['K30'] = int(self.dsm_cl_cm.get())
                self.bill_sheet['K32'] = int(self.tech_cl_cm.get())
                self.bill_sheet['K34'] = int(self.mgr_cl_cm.get())
                self.bill_sheet['K36'] = int(self.dsm_fs_cm.get())
                self.bill_sheet['K38'] = int(self.tech_fs_cm.get())
                self.bill_sheet['K40'] = int(self.mgr_fs_cm.get())

                self.bill_sheet['K13'] = int(self.dsm_allw_cm.get())
                self.bill_sheet['K15'] = int(self.tech_allw_cm.get())
                self.bill_sheet['K17'] = int(self.mgr_allw_cm.get())
                self.bill_sheet['K19'] = int(self.dsm_fh_allw_cm.get())
                self.bill_sheet['K21'] = int(self.tech_fh_allw_cm.get())
                self.bill_sheet['K23'] = int(self.mgr_fh_allw_cm.get())
                self.bill_sheet['K25'] = int(self.dsm_nh_allw_cm.get())
                self.bill_sheet['K27'] = int(self.tech_nh_allw_cm.get())
                self.bill_sheet['K29'] = int(self.mgr_nh_allw_cm.get())
                self.bill_sheet['K31'] = int(self.dsm_cl_allw_cm.get())
                self.bill_sheet['K33'] = int(self.tech_cl_allw_cm.get())
                self.bill_sheet['K35'] = int(self.mgr_cl_allw_cm.get())
                self.bill_sheet['K37'] = int(self.dsm_fs_allw_cm.get())
                self.bill_sheet['K39'] = int(self.tech_fs_allw_cm.get())
                self.bill_sheet['K41'] = int(self.mgr_fs_allw_cm.get())

                self.billWorkbook.save(self.billPath)
            
                # current_month_claimed_mandays_df = pd.read_excel(self.current_month_claimed_mandays)
                # last_month_claimed_mandays_df = pd.read_excel(self.last_month_claimed_mandays)
                current_month_active_mandays_df = pd.read_excel(self.current_month_active_mandays)

                if self.filePath is not None:
                    wage_rate_df = pd.read_excel(self.filePath)
                else:
                    wage_rate_df = pd.read_excel(util.get_wage_template()) 

                generateBill.createBill(self.billPath, self.current_month_claimed_mandays, self.last_month_claimed_mandays, current_month_active_mandays_df, wage_rate_df, self.lastMonth, self.billMonth, self.lastYear, self.billYear)           

                messagebox.showinfo("Success", "Bill Generated successfully")

                win = Toplevel()
                dashboard.Dashboard(win)
                self.window.withdraw()
                win.deiconify()

            else:
                # Display an error message if any field is empty
                messagebox.showerror("Error", "Please fill in all the fields.")
                
        except Exception as e:
            print(e)

    def back_operation(self):
        win = Toplevel()
        billView.BillView(win)
        self.window.withdraw()
        win.deiconify()






        


