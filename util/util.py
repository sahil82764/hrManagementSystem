import os, sys
import datetime
from openpyxl import load_workbook
from tkinter import messagebox

def create_or_open_folder(fPath):
    if os.path.exists(fPath):
        return fPath
    else:
        os.makedirs(fPath)
        return fPath

def get_custom_template(type):
    if type == 'Mandays':
        return 'customTemplate\\mandaysClaimed\\customTemplateMandays.xlsx'
    elif type == 'Bill':
        return 'customTemplate\\vendorBill\\customTemplateBill.xlsx'
    else:
        print('NO FILE FOUND')

def save_mandays(type, year, month, vName, sName):
    typefolder = f'Mandays\{type}\{year}\{month}'
    typePath = create_or_open_folder(typefolder)
    return f'{typePath}\{vName} - {sName}.xlsx'

def save_bill(year, month, vName, sName):
    billfolder = f'Bills\{year}\{month}'
    billPath = create_or_open_folder(billfolder)
    if sName == "":
        return f'{billPath}\{vName}.xlsx'    
    else:
        return f'{billPath}\{vName} - {sName}.xlsx'

def get_mandays(year, month, vName, sName):
    return f'Mandays\\Active\\{year}\\{month}\\{vName} - {sName}.xlsx'

def get_wage_template():
    return f'customTemplate\\wageRate\\customWageRate.xlsx'

def get_estimate_path(year, month, vName, sName):
    try:
        destination_folder = os.path.dirname(f'salaryEstimate\\{year}\\{month}\\{vName} - {sName}.xlsx')
        
        if not os.path.exists(destination_folder):
            os.makedirs(destination_folder)
            return f'salaryEstimate\\{year}\\{month}\\{vName} - {sName}.xlsx'
        else:
            return f'salaryEstimate\\{year}\\{month}\\{vName} - {sName}.xlsx'
    except Exception as e:
            print(f"An error occurred: {e} at line {sys.exc_info()[-1].tb_lineno}")

def get_attendance_path(year, month, vName, sName):
    try:
        destination_folder = os.path.dirname(f'attendanceRecord\\{year}\\{month}\\{vName} - {sName}.xlsx')
        
        if not os.path.exists(destination_folder):
            os.makedirs(destination_folder)
            return f'attendanceRecord\\{year}\\{month}\\{vName} - {sName}.xlsx'
        else:
            return f'attendanceRecord\\{year}\\{month}\\{vName} - {sName}.xlsx'
    except Exception as e:
            print(f"An error occurred: {e} at line {sys.exc_info()[-1].tb_lineno}")

def count_days(year, month):
    weekdays = 0
    sundays = 0

    # Get the first day of the month
    first_day = datetime.date(year, month, 1)

    # Get the last day of the month
    if month == 12:
        last_day = datetime.date(year + 1, 1, 1)
    else:
        last_day = datetime.date(year, month + 1, 1)

    # Iterate through the days of the month
    current_day = first_day
    while current_day < last_day:
        # 0 - Monday, 6 - Sunday
        if current_day.weekday() == 6:
            sundays += 1
        elif current_day.weekday() <= 5:
            weekdays += 1
        current_day += datetime.timedelta(days=1)

    return weekdays, sundays

def get_spi_claim(nonBusSale, busSale, mandays, selectedStation):
    
    spi = (nonBusSale + (busSale/2))/mandays
    
    if "DTC" in selectedStation:
        return 0
    else:    
        if spi <= 300:
            return 0
        elif spi >= 301 and spi <= 350:
            return 7000
        elif spi >= 351 and spi <= 400:
            return 9000
        elif spi >= 401 and spi <= 450:
            return 11000
        elif spi >= 451 and spi <= 500:
            return 13000
        else:
            return 15000
        
def find_common_sequence(strings):
    words_list = [s.split() for s in strings]
    common_sequence = []

    for word_group in zip(*words_list):
        if all(w == word_group[0] for w in word_group):
            common_sequence.append(word_group[0])
        else:
            break

    return " ".join(common_sequence)

def combine_names(names):
    common_sequence = find_common_sequence(names)
    suffixes = [name.replace(common_sequence, "").strip() for name in names]

    output = f"{common_sequence} {' & '.join(suffixes)}"
    return output

def mergeBill2(file1, file2):
    outputFile = get_custom_template('Bill')

    file1Name = os.path.basename(file1)
    file2Name = os.path.basename(file2)
    
    if file1.split("\\")[-2] == file2.split("\\")[-2]:
        if file1Name.split("-")[0].strip() != file2Name.split("-")[0].strip():
            messagebox.showerror("ERROR", "PLEASE ATTACH SAME VENDOR/MONTH FILE.")
        else:
            try:
                outputWorkbook = load_workbook(outputFile)
                output_sheet = outputWorkbook.active

                file1Workbook = load_workbook(file1)
                file1_sheet = file1Workbook.active

                file2Workbook = load_workbook(file2)
                file2_sheet = file2Workbook.active

                # Sheet intro info
                output_sheet['B2'] = file1_sheet['B2'].value
                output_sheet['B3'] = file1_sheet['B3'].value
                output_sheet['B4'] = file1_sheet['B4'].value
                output_sheet['B5'] = combine_names([file1_sheet['B5'].value, file2_sheet['B5'].value])
                output_sheet['B6'] = file1_sheet['B6'].value
                output_sheet['G2'] = file1_sheet['G2'].value
                output_sheet['G3'] = file1_sheet['G3'].value
                output_sheet['G4'] = file1_sheet['G4'].value
                output_sheet['G5'] = file1_sheet['G5'].value
                output_sheet['G6'] = file1_sheet['G6'].value

                #PreviousMonth MANDAYS CLAIMED
                for row in range(12,42):
                    output_sheet[f'C{row}'] = file1_sheet[f'C{row}'].value + file2_sheet[f'C{row}'].value

                #PreviousMonth MANDAYS DISBURSED
                for row in range(12,42):
                    output_sheet[f'D{row}'] = file1_sheet[f'D{row}'].value + file2_sheet[f'D{row}'].value

                #PreviousMonth WEEKOFF
                for row in range(12,42):
                    output_sheet[f'E{row}'] = file1_sheet[f'E{row}'].value + file2_sheet[f'E{row}'].value

                #CurrentMonth MANDAYS
                for row in range(12,42):
                    output_sheet[f'H{row}'] = file1_sheet[f'H{row}'].value + file2_sheet[f'H{row}'].value

                # WAGES
                for row in range(12,42):
                    output_sheet[f'J{row}'] = file1_sheet[f'J{row}'].value
                    output_sheet[f'K{row}'] = file1_sheet[f'K{row}'].value

                # I42-I46
                for row in range(42,47):
                    output_sheet[f'I{row}'] = file1_sheet[f'I{row}'].value + file2_sheet[f'I{row}'].value

                # LUMPSUM REIMBURSEMENT
                for row in range(50,53):
                    output_sheet[f'H{row}'] = (0 if file1_sheet[f'H{row}'].value is None else file1_sheet[f'H{row}'].value) + (0 if file2_sheet[f'H{row}'].value is None else file2_sheet[f'H{row}'].value)

                # OTHER EXPENSES REIMBURSEMENT
                for row in range(54,62):
                    output_sheet[f'H{row}'] = (0 if file1_sheet[f'H{row}'].value is None else file1_sheet[f'H{row}'].value) + (0 if file2_sheet[f'H{row}'].value is None else file2_sheet[f'H{row}'].value)

                # OPERATOR CHARGES
                for row in range(63,65):
                    output_sheet[f'H{row}'] = 0 if file1_sheet[f'H{row}'].value is None else file1_sheet[f'H{row}'].value + 0 if file2_sheet[f'H{row}'].value is None else file2_sheet[f'H{row}'].value

                # MANPOWER DETAILS                
                for row in range(8,10):
                    output_sheet[f'F{row}'] = 0 if file1_sheet[f'F{row}'].value is None else file1_sheet[f'F{row}'].value + 0 if file2_sheet[f'F{row}'].value is None else file2_sheet[f'F{row}'].value
                    output_sheet[f'G{row}'] = 0 if file1_sheet[f'G{row}'].value is None else file1_sheet[f'G{row}'].value + 0 if file2_sheet[f'G{row}'].value is None else file2_sheet[f'G{row}'].value
                    output_sheet[f'H{row}'] = 0 if file1_sheet[f'H{row}'].value is None else file1_sheet[f'H{row}'].value + 0 if file2_sheet[f'H{row}'].value is None else file2_sheet[f'H{row}'].value

                #DAYS AND MONTHS ENTRIES
                output_sheet['M1'] = file1_sheet['M1'].value
                output_sheet['T1'] = file1_sheet['T1'].value

                output_sheet['M2'] = file1_sheet['M2'].value
                output_sheet['T2'] = file1_sheet['T2'].value

                output_sheet['M3'] = file1_sheet['M3'].value
                output_sheet['P3'] = file1_sheet['P3'].value

                output_sheet['T3'] = file1_sheet['T3'].value
                output_sheet['W3'] = file1_sheet['W3'].value



                outputPath = save_bill(file1.split("\\")[-3], file1.split("\\")[-2], file1.split("\\")[-1].split("-")[0].strip(), "" )

                outputWorkbook.save(outputPath)
                messagebox.showinfo("Success", "Bill Generated successfully")

            except Exception as e:
                print(f"An error occurred: {e} at line {sys.exc_info()[-1].tb_lineno}")