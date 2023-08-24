from openpyxl import load_workbook
import calendar
from util import util



def createBill(billPath, current_month_claimed_mandays, last_month_claimed_mandays, current_month_active_mandays_df, wage_rate_df, lastMonth, billMonth, lastYear, billYear):

    billWOrkbook = load_workbook(billPath)
    active_bill_sheet = billWOrkbook.active

    last_month_estimate = load_workbook(last_month_claimed_mandays)
    last_month_estimate_sheet = last_month_estimate.active

    current_month_estimate = load_workbook(current_month_claimed_mandays)
    current_month_estimate_sheet = current_month_estimate.active

    # =============== Last Month Mandays Claimed cell-C11 ===============

    active_bill_sheet['C12'] = last_month_estimate_sheet['C8'].value
    active_bill_sheet['C13'] = last_month_estimate_sheet['C8'].value
    active_bill_sheet['C14'] = last_month_estimate_sheet['C7'].value
    active_bill_sheet['C15'] = last_month_estimate_sheet['C7'].value
    active_bill_sheet['C16'] = last_month_estimate_sheet['C6'].value
    active_bill_sheet['C17'] = last_month_estimate_sheet['C6'].value

    active_bill_sheet['C18'] = last_month_estimate_sheet['E8'].value
    active_bill_sheet['C19'] = last_month_estimate_sheet['E8'].value
    active_bill_sheet['C20'] = last_month_estimate_sheet['E7'].value
    active_bill_sheet['C21'] = last_month_estimate_sheet['E7'].value
    active_bill_sheet['C22'] = last_month_estimate_sheet['E6'].value
    active_bill_sheet['C23'] = last_month_estimate_sheet['E6'].value

    active_bill_sheet['C24'] = last_month_estimate_sheet['F8'].value
    active_bill_sheet['C25'] = last_month_estimate_sheet['F8'].value
    active_bill_sheet['C26'] = last_month_estimate_sheet['F7'].value
    active_bill_sheet['C27'] = last_month_estimate_sheet['F7'].value
    active_bill_sheet['C28'] = last_month_estimate_sheet['F6'].value
    active_bill_sheet['C29'] = last_month_estimate_sheet['F6'].value

    active_bill_sheet['C30'] = last_month_estimate_sheet['D8'].value
    active_bill_sheet['C31'] = last_month_estimate_sheet['D8'].value
    active_bill_sheet['C32'] = last_month_estimate_sheet['D7'].value
    active_bill_sheet['C33'] = last_month_estimate_sheet['D7'].value
    active_bill_sheet['C34'] = last_month_estimate_sheet['D6'].value
    active_bill_sheet['C35'] = last_month_estimate_sheet['D6'].value

    active_bill_sheet['C36'] = last_month_estimate_sheet['G8'].value
    active_bill_sheet['C37'] = last_month_estimate_sheet['G8'].value
    active_bill_sheet['C38'] = last_month_estimate_sheet['G7'].value
    active_bill_sheet['C39'] = last_month_estimate_sheet['G7'].value
    active_bill_sheet['C40'] = last_month_estimate_sheet['G6'].value
    active_bill_sheet['C41'] = last_month_estimate_sheet['G6'].value

    # =============== Current Month Mandays Claimed cell-H11 ===============

    active_bill_sheet['H12'] = current_month_estimate_sheet['C8'].value
    active_bill_sheet['H13'] = current_month_estimate_sheet['C8'].value
    active_bill_sheet['H14'] = current_month_estimate_sheet['C7'].value
    active_bill_sheet['H15'] = current_month_estimate_sheet['C7'].value
    active_bill_sheet['H16'] = current_month_estimate_sheet['C6'].value
    active_bill_sheet['H17'] = current_month_estimate_sheet['C6'].value

    active_bill_sheet['H18'] = current_month_estimate_sheet['E8'].value
    active_bill_sheet['H19'] = current_month_estimate_sheet['E8'].value
    active_bill_sheet['H20'] = current_month_estimate_sheet['E7'].value
    active_bill_sheet['H21'] = current_month_estimate_sheet['E7'].value
    active_bill_sheet['H22'] = current_month_estimate_sheet['E6'].value
    active_bill_sheet['H23'] = current_month_estimate_sheet['E6'].value

    active_bill_sheet['H24'] = current_month_estimate_sheet['F8'].value
    active_bill_sheet['H25'] = current_month_estimate_sheet['F8'].value
    active_bill_sheet['H26'] = current_month_estimate_sheet['F7'].value
    active_bill_sheet['H27'] = current_month_estimate_sheet['F7'].value
    active_bill_sheet['H28'] = current_month_estimate_sheet['F6'].value
    active_bill_sheet['H29'] = current_month_estimate_sheet['F6'].value

    active_bill_sheet['H30'] = current_month_estimate_sheet['D8'].value
    active_bill_sheet['H31'] = current_month_estimate_sheet['D8'].value
    active_bill_sheet['H32'] = current_month_estimate_sheet['D7'].value
    active_bill_sheet['H33'] = current_month_estimate_sheet['D7'].value
    active_bill_sheet['H34'] = current_month_estimate_sheet['D6'].value
    active_bill_sheet['H35'] = current_month_estimate_sheet['D6'].value

    active_bill_sheet['H36'] = current_month_estimate_sheet['G8'].value
    active_bill_sheet['H37'] = current_month_estimate_sheet['G8'].value
    active_bill_sheet['H38'] = current_month_estimate_sheet['G7'].value
    active_bill_sheet['H39'] = current_month_estimate_sheet['G7'].value
    active_bill_sheet['H40'] = current_month_estimate_sheet['G6'].value
    active_bill_sheet['H41'] = current_month_estimate_sheet['G6'].value

    # =============== Current Month Mandays Disbursed cell-D11 ===============

    active_bill_sheet['D12'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'DSM', 'WD'].values[0]
    active_bill_sheet['D13'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'DSM', 'WD'].values[0]
    active_bill_sheet['D14'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'TECH', 'WD'].values[0]
    active_bill_sheet['D15'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'TECH', 'WD'].values[0]
    active_bill_sheet['D16'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'MGR', 'WD'].values[0]
    active_bill_sheet['D17'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'MGR', 'WD'].values[0]

    active_bill_sheet['D18'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'DSM', 'FH'].values[0]
    active_bill_sheet['D19'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'DSM', 'FH'].values[0]
    active_bill_sheet['D20'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'TECH', 'FH'].values[0]
    active_bill_sheet['D21'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'TECH', 'FH'].values[0]
    active_bill_sheet['D22'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'MGR', 'FH'].values[0]
    active_bill_sheet['D23'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'MGR', 'FH'].values[0]

    active_bill_sheet['D24'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'DSM', 'NH'].values[0]
    active_bill_sheet['D25'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'DSM', 'NH'].values[0]
    active_bill_sheet['D26'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'TECH', 'NH'].values[0]
    active_bill_sheet['D27'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'TECH', 'NH'].values[0]
    active_bill_sheet['D28'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'MGR', 'NH'].values[0]
    active_bill_sheet['D29'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'MGR', 'NH'].values[0]

    active_bill_sheet['D30'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'DSM', 'CL'].values[0]
    active_bill_sheet['D31'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'DSM', 'CL'].values[0]
    active_bill_sheet['D32'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'TECH', 'CL'].values[0]
    active_bill_sheet['D33'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'TECH', 'CL'].values[0]
    active_bill_sheet['D34'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'MGR', 'CL'].values[0]
    active_bill_sheet['D35'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'MGR', 'CL'].values[0]

    active_bill_sheet['D36'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'DSM', 'FT'].values[0]
    active_bill_sheet['D37'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'DSM', 'FT'].values[0]
    active_bill_sheet['D38'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'TECH', 'FT'].values[0]
    active_bill_sheet['D39'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'TECH', 'FT'].values[0]
    active_bill_sheet['D40'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'MGR', 'FT'].values[0]
    active_bill_sheet['D41'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'MGR', 'FT'].values[0]

    # =============== Current Month Week OFF cell-E11 ===============

    active_bill_sheet['E12'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'DSM', 'OFF'].values[0]
    active_bill_sheet['E13'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'DSM', 'OFF'].values[0]
    active_bill_sheet['E14'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'TECH', 'OFF'].values[0]
    active_bill_sheet['E15'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'TECH', 'OFF'].values[0]
    active_bill_sheet['E16'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'MGR', 'OFF'].values[0]
    active_bill_sheet['E17'] = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'MGR', 'OFF'].values[0]

    # =============== Last Month Reconciliation cell-F11  ===============

    # for row in range(12,42):
    #     c_cell = f'C{row}'
    #     d_cell = f'D{row}'
    #     f_cell = f'F{row}'
    #     e_cell = f'E{row}'
        
    #     d_value = active_bill_sheet[d_cell].value
    #     c_value = active_bill_sheet[c_cell].value
    #     e_value = active_bill_sheet[e_cell].value

    #     if row>13 and row<=17:        
    #         if d_value is not None and c_value is not None:
    #             active_bill_sheet[f_cell].value = d_value + e_value - c_value
    #     else:
    #         if d_value is not None and c_value is not None:
    #             active_bill_sheet[f_cell].value = d_value - c_value

    # =============== Last Month Amount cell-G11  ===============

    # for row in range(12,42):   
    #     f_cell = f'F{row}'
    #     b_cell = f'B{row}'
    #     g_cell = f'G{row}'
            
    #     recon_value = active_bill_sheet[f_cell].value
    #     filter_value = active_bill_sheet['B10'].value
    #     particular = active_bill_sheet[b_cell].value
    #     wage = wage_rate_df.loc[wage_rate_df[filter_value] == particular, 'Wage_Rate_1'].values[0]
    #     days = calendar.monthrange(lastYear, lastMonth)[1]

    #     if (row>17 and row<=29) or (row>35 and row<=41): 
    #         active_bill_sheet[g_cell].value = round(( recon_value * wage ) / 26, 0)
    #     else:
    #         active_bill_sheet[g_cell].value = round(( recon_value * wage ) / days, 0)

    # =============== Current Month Amount cell-I11  ===============

    # for row in range(12,42):   
    #     h_cell = f'H{row}'
    #     b_cell = f'B{row}'
    #     i_cell = f'I{row}'
            
    #     mandays_value = active_bill_sheet[h_cell].value
    #     filter_value = active_bill_sheet['B10'].value
    #     particular = active_bill_sheet[b_cell].value
    #     wage = wage_rate_df.loc[wage_rate_df[filter_value] == particular, 'Wage_Rate_2'].values[0]
    #     days = calendar.monthrange(billYear, billMonth)[1]

    #     if (row>17 and row<=29) or (row>35 and row<=41): 
    #         active_bill_sheet[i_cell].value = round(( mandays_value * wage ) / 26, 0)
    #     else:
    #         active_bill_sheet[i_cell].value = round(( mandays_value * wage ) / days, 0)

    # =============== CELL: M1 and T1  ===============
    active_bill_sheet['M1'] = calendar.monthrange(billYear, billMonth)[1]
    active_bill_sheet['T1'] = calendar.monthrange(lastYear, lastMonth)[1]

    active_bill_sheet['M2'] = f"{calendar.month_abbr[billMonth]}-{billYear}"
    active_bill_sheet['T2'] = f"{calendar.month_abbr[lastMonth]}-{lastYear}"

    active_bill_sheet['M3'], active_bill_sheet['P3'] = util.count_days(billYear, billMonth)
    active_bill_sheet['T3'], active_bill_sheet['W3'] = util.count_days(lastYear, lastMonth)

    # =============== LUMPSUM REIMBURSEMENT CELL: H50-HH53 ===============
    active_bill_sheet['H50'] = current_month_estimate_sheet['C14'].value
    active_bill_sheet['H51'] = current_month_estimate_sheet['C15'].value
    active_bill_sheet['H52'] = current_month_estimate_sheet['C16'].value

    # =============== OTHER EXPENSES REIMBURSEMENT CELL: H55-H64 =========
    active_bill_sheet['H54'] = current_month_estimate_sheet['C20'].value
    active_bill_sheet['H55'] = current_month_estimate_sheet['C21'].value
    active_bill_sheet['H56'] = current_month_estimate_sheet['C22'].value
    active_bill_sheet['H57'] = current_month_estimate_sheet['C23'].value
    active_bill_sheet['H58'] = current_month_estimate_sheet['C24'].value
    active_bill_sheet['H59'] = current_month_estimate_sheet['C25'].value
    active_bill_sheet['H60'] = current_month_estimate_sheet['C26'].value
    active_bill_sheet['H61'] = current_month_estimate_sheet['C27'].value

    # =============== OPERATOR SERVICE CHARGES CELL: H66-H67 ==============
    nonBusSale = current_month_estimate_sheet['G31'].value
    busSale = current_month_estimate_sheet['D31'].value
    mandays = current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'DSM', 'WD'].values[0] + current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'TECH', 'WD'].values[0] + current_month_active_mandays_df.loc[current_month_active_mandays_df['Designation'] == 'MGR', 'WD'].values[0]
    
    active_bill_sheet['H63'] = 0
    selectedStation = active_bill_sheet['B5'].value
    active_bill_sheet['H64'] = util.get_spi_claim(nonBusSale, busSale, mandays, selectedStation)

    # =============== MANPOWER DEPLOYED CELLS =============================
    active_bill_sheet['F8'] = current_month_estimate_sheet['I22'].value
    active_bill_sheet['F9'] = current_month_estimate_sheet['F16'].value
    active_bill_sheet['G8'] = current_month_estimate_sheet['I21'].value
    active_bill_sheet['G9'] = current_month_estimate_sheet['F15'].value
    active_bill_sheet['H8'] = current_month_estimate_sheet['I20'].value
    active_bill_sheet['H9'] = current_month_estimate_sheet['F14'].value


    # =============== Saving Workbook  ===============
    billWOrkbook.save(billPath)




